"""Map yellow input cells in CBAM template for fill mapping."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)


def is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return False
    rgb = str(getattr(fill.fgColor, "rgb", "") or "")
    return rgb.upper().endswith("FFFF00") or rgb.upper() == "FFFFFF00"


def nearby_label(ws, r: int, c: int) -> str:
    for dc in range(1, 6):
        left = ws.cell(r, max(1, c - dc)).value
        if isinstance(left, str) and left.startswith("="):
            # try Translations reference - just show formula
            return left[:60]
        if left is not None and not (isinstance(left, str) and left.startswith("=")):
            return str(left)[:60]
    up = ws.cell(max(1, r - 1), c).value
    if up is not None:
        return str(up)[:60]
    return ""


for name in ["A_InstData", "Summary_Products", "Summary_Processes", "D_Processes", "E_PurchPrec", "B_EmInst", "C_Emissions&Energy"]:
    ws = wb[name]
    print(f"\n==== {name} yellow inputs (first 80) ====")
    n = 0
    for row in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row or 1), max_col=min(30, ws.max_column or 1)):
        for cell in row:
            if not is_yellow(cell):
                continue
            # skip if already has formula
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            label = nearby_label(ws, cell.row, cell.column)
            print(f"{cell.coordinate}\tval={cell.value!r}\tnear={label!r}")
            n += 1
            if n >= 80:
                break
        if n >= 80:
            break
    print(f"... total yellow non-formula shown {n}")
