"""Inspect EU CBAM template labels and yellow input cells."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)
tr = wb["Translations"]

for r in [64, 65, 67, 68, 69, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 89, 90, 91, 92, 93]:
    print(f"T{r}: {tr[f'C{r}'].value!r}")


def is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return False
    for attr in ("fgColor", "start_color"):
        color = getattr(fill, attr, None)
        if color is None or not getattr(color, "rgb", None):
            continue
        c = str(color.rgb).upper()
        if "FFFF00" in c:
            return True
    return False


print("\n=== A_InstData rows 9-50 ===")
ws = wb["A_InstData"]
for r in range(9, 50):
    for col in list("EFGHIJKLM"):
        cell = ws[f"{col}{r}"]
        y = is_yellow(cell)
        if y or (cell.value is not None and col in "IL"):
            print(f"{cell.coordinate}: yellow={y} val={cell.value!r}"[:140])

print("\n=== Summary_Products rows 7-12 (yellow or non-formula) ===")
sp = wb["Summary_Products"]
for r in range(7, 12):
    for c in range(1, 45):
        cell = sp.cell(r, c)
        y = is_yellow(cell)
        if not y and cell.value is None:
            continue
        v = cell.value
        if isinstance(v, str) and v.startswith("=") and not y:
            continue
        print(f"{cell.coordinate}: y={y} v={v!r}"[:140])

print("\n=== Summary_Products row 8/9 headers (all with values) ===")
for r in (8, 9):
    for c in range(1, 45):
        cell = sp.cell(r, c)
        if cell.value is not None:
            print(f"{cell.coordinate}: {cell.value!r}"[:140])
