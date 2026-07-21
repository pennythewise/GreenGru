"""Inspect derived process names and InputOutput SEE rows."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)

ws = wb["A_InstData"]
print("A_InstData process helper cols S/T/E/F/L row 83:")
for col in "DEFGHIJKLMNOST":
    cell = ws[f"{col}83"]
    print(f"  {cell.coordinate}: {cell.value!r}")

print("\nCONST_LIST_Goods values:")
pc = wb["Parameters_Constants"]
for r in range(89, 107):
    print(f"  A{r}: {pc[f'A{r}'].value!r}")

print("\nInputOutput D71:D80 and AK/AM/AO:")
io = wb["InputOutput"]
for r in range(65, 85):
    vals = []
    for col in ["D", "AK", "AM", "AO"]:
        cell = io[f"{col}{r}"]
        if cell.value is not None:
            vals.append(f"{cell.coordinate}={cell.value!r}"[:90])
    if vals:
        print(f"R{r}: " + " | ".join(vals))

# CN codes containing 7318
pcn = wb["Parameters_CNCodes"]
print("\nCN codes with 7318:")
for r in range(4, 573):
    v = pcn[f"D{r}"].value
    if v and "7318" in str(v).replace(" ", ""):
        print(f"  D{r}={v!r} E={pcn[f'E{r}'].value!r}")
