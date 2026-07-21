"""Inspect Summary_Products row 10 formulas and A_InstData goods/process yellow cells."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)
sp = wb["Summary_Products"]
tr = wb["Translations"]

print("Summary_Products row 10 all columns A-AR:")
for c in range(1, 45):
    cell = sp.cell(10, c)
    if cell.value is not None or (cell.fill and cell.fill.fgColor and getattr(cell.fill.fgColor, "rgb", None)):
        rgb = getattr(getattr(cell.fill, "fgColor", None), "rgb", None)
        print(f"  {cell.coordinate}: rgb={rgb} val={cell.value!r}"[:160])

print("\nA_InstData goods/routes area 55-95:")
ws = wb["A_InstData"]
for r in range(55, 95):
    vals = []
    for c in range(3, 15):
        cell = ws.cell(r, c)
        if cell.value is not None:
            vals.append(f"{cell.coordinate}={cell.value!r}"[:70])
    if vals:
        print(f"R{r}: " + " | ".join(vals[:6]))

print("\nTranslations for goods section:")
for r in range(100, 130):
    v = tr[f"C{r}"].value
    if v:
        print(f"  C{r}: {v!r}")

print("\nSummary_Processes structure rows 1-30:")
spr = wb["Summary_Processes"]
for r in range(1, 35):
    vals = []
    for c in range(1, 20):
        cell = spr.cell(r, c)
        if cell.value is not None:
            vals.append(f"{cell.coordinate}={cell.value!r}"[:80])
    if vals:
        print(f"R{r}: " + " | ".join(vals[:5]))
