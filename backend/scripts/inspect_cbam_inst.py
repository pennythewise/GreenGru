"""Deep inspect A_InstData installation inputs and Summary_Products columns."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)
ws = wb["A_InstData"]
tr = wb["Translations"]


def fill_info(cell):
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return "nofill"
    parts = [f"type={fill.fill_type}"]
    for attr in ("fgColor", "bgColor", "start_color", "end_color"):
        color = getattr(fill, attr, None)
        if color is None:
            continue
        parts.append(f"{attr}=type:{color.type},rgb:{getattr(color, 'rgb', None)},theme:{getattr(color, 'theme', None)},indexed:{getattr(color, 'indexed', None)}")
    return " | ".join(parts)


print("Merged ranges involving I19:I45:")
for mr in ws.merged_cells.ranges:
    s = str(mr)
    if any(f"I{r}" in s or f"L{r}" in s for r in range(9, 50)):
        print(" ", s)

print("\nRows 19-45 col I detail:")
for r in range(19, 46):
    cell = ws[f"I{r}"]
    e = ws[f"E{r}"].value
    # resolve translation label if possible
    label = None
    if isinstance(e, str) and "Translations" in e:
        # =Translations!$B$75
        try:
            ref = e.split("$B$")[1].rstrip(")")
            label = tr[f"C{ref}"].value
        except Exception:
            label = e
    print(f"I{r}: val={cell.value!r} fill={fill_info(cell)}")
    print(f"     E label: {label!r}")

print("\nSummary_Products EN headers from Translations C:")
for r in [472, 474, 483, 485, 570, 574, 603, 604, 605, 606, 607, 608, 610, 611, 612, 613, 614, 615]:
    print(f"  C{r}: {tr[f'C{r}'].value!r}")

print("\nSummary_Communication yellow non-formula (sample):")
sc = wb["Summary_Communication"]


def is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return False
    for attr in ("fgColor", "start_color"):
        color = getattr(fill, attr, None)
        if color is None or not getattr(color, "rgb", None):
            continue
        if "FFFF00" in str(color.rgb).upper():
            return True
    return False


count = 0
for row in sc.iter_rows(min_row=1, max_row=120, max_col=40):
    for cell in row:
        if not is_yellow(cell):
            continue
        if isinstance(cell.value, str) and cell.value.startswith("="):
            continue
        print(f"  {cell.coordinate}: {cell.value!r}")
        count += 1
        if count >= 40:
            break
    if count >= 40:
        break

print("\nD_Processes key yellow near top:")
dp = wb["D_Processes"]
for r in range(1, 80):
    for c in range(1, 20):
        cell = dp.cell(r, c)
        if is_yellow(cell) and not (isinstance(cell.value, str) and str(cell.value).startswith("=")):
            print(f"  {cell.coordinate}: {cell.value!r}")
