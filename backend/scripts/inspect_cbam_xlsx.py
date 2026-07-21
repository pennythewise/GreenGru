"""Inspect CBAM Communication template input cells."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)


def cell_color(cell) -> str:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return ""
    rgb = getattr(fill.fgColor, "rgb", None)
    if rgb and str(rgb) not in ("00000000", "0"):
        return str(rgb)
    theme = getattr(fill.fgColor, "theme", None)
    if theme is not None:
        return f"theme{theme}"
    return ""


def scan_sheet(name: str, max_r: int = 140, max_c: int = 18) -> None:
    ws = wb[name]
    print(f"\n######## {name} ########")
    for r in range(1, min(max_r, (ws.max_row or 1) + 1)):
        bits: list[str] = []
        for c in range(1, min(max_c, (ws.max_column or 1) + 1)):
            cell = ws.cell(r, c)
            color = cell_color(cell)
            val = cell.value
            if val is None and not color:
                continue
            kind = "F" if isinstance(val, str) and str(val).startswith("=") else ("V" if val is not None else "E")
            # Prefer showing non-formula values and colored empties (likely inputs)
            if kind == "F" and not color:
                continue
            s = repr(val)[:50] if val is not None else ""
            bits.append(f"{cell.coordinate}:{kind}:{color[-8:] if color else '-'}:{s}")
        if bits:
            line = " | ".join(bits)
            print(f"R{r}: {line[:260]}")


for sn in [
    "A_InstData",
    "Summary_Communication",
    "Summary_Products",
    "Summary_Processes",
    "D_Processes",
    "E_PurchPrec",
    "B_EmInst",
]:
    scan_sheet(sn)

# Also dump data_only=False non-formula values in A_InstData that look like placeholders
print("\n######## A_InstData NON-FORMULA VALUES ########")
ws = wb["A_InstData"]
for row in ws.iter_rows(min_row=1, max_row=min(160, ws.max_row or 1), max_col=min(20, ws.max_column or 1)):
    for cell in row:
        v = cell.value
        if v is None:
            continue
        if isinstance(v, str) and v.startswith("="):
            continue
        print(f"{cell.coordinate}={v!r}")
