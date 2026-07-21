"""Find valid dropdown list values for goods/process/CN codes."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedNameDict

path = Path(r"c:\Users\ACER\Downloads\CBAM Communication template for installations_en_20241213.xlsx")
wb = load_workbook(path, data_only=False)

# Sample named ranges related to goods / CN
names = []
for dn in wb.defined_names.values():
    n = dn.name
    if any(k in n.lower() for k in ("good", "prod", "cncode", "cn_", "route", "exist")):
        names.append(n)
print("Relevant defined names:", sorted(names)[:80])

# Try to resolve a few
for name in [
    "CNTR_List_ExistProdProcNames",
    "CNTR_List_ExistProdProc",
    "CONST_LIST_Goods",
    "CNCodes_ListKey",
]:
    try:
        defn = wb.defined_names[name]
        print(f"\n{name}: {defn.attr_text}")
    except KeyError:
        print(f"\n{name}: NOT FOUND")

# Look at c_CodeLists for Iron or steel / CN codes around 7318
cl = wb["c_CodeLists"]
print("\nc_CodeLists sample for steel / 7318:")
hits = 0
for row in cl.iter_rows(min_row=1, max_row=min(cl.max_row, 800), max_col=8):
    for cell in row:
        if cell.value and ("7318" in str(cell.value) or "Iron or steel" in str(cell.value) or "iron or steel" in str(cell.value).lower()):
            print(f"  {cell.coordinate}: {cell.value!r}")
            hits += 1
            if hits > 30:
                break
    if hits > 30:
        break

# Data validation on E62 and D10
ws = wb["A_InstData"]
print("\nData validations on A_InstData:")
if ws.data_validations:
    for dv in ws.data_validations.dataValidation:
        sq = str(dv.sqref)
        if any(x in sq for x in ("E62", "E83", "G62", "I62", "L83")):
            print(f"  {sq}: type={dv.type} formula1={dv.formula1!r}")

sp = wb["Summary_Products"]
print("\nData validations on Summary_Products D10/F10:")
if sp.data_validations:
    for dv in sp.data_validations.dataValidation:
        sq = str(dv.sqref)
        if "D10" in sq or "F10" in sq or "$D$10" in sq or "D10:" in sq:
            print(f"  {sq}: type={dv.type} formula1={dv.formula1!r}")
