"""Smoke tests for filled EU CBAM Communication template export."""

from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.services.cbam_communication_xlsx import (
    fill_cbam_communication_xlsx,
    template_path,
)


DEMO = {
    "installation_name_en": "Ningbo Hengfeng Precision Fasteners Co., Ltd.",
    "installation_name_optional": "宁波恒峰精密紧固件有限公司",
    "street_number": "No. 88 Jingang Road, Beilun District",
    "economic_activity": "Manufacture of fasteners and structural steel components",
    "country": "CN",
    "city": "Ningbo",
    "post_code": "315800",
    "unlocode": "CNNGB",
    "latitude": "29.8683",
    "longitude": "121.5440",
    "period_start": "2025-01-01",
    "period_end": "2025-12-31",
    "aggregated_goods_category": "Iron or steel products",
    "process_name": "Bolt & fastener finishing line",
    "cn_code": "7318 15 88",
    "product_name": "Hex bolt M12",
    "see_direct": "1.87",
    "see_indirect": "0.12",
    "scrap_per_t_steel": "0.245",
}


def test_template_present():
    assert template_path().is_file()


def test_fill_writes_instdata_and_products():
    out = fill_cbam_communication_xlsx(DEMO)
    assert out.is_file()
    assert out.stat().st_size > 100_000

    wb = load_workbook(out, data_only=False)
    try:
        assert set(wb.sheetnames) >= {
            "A_InstData",
            "Summary_Communication",
            "Summary_Products",
            "Summary_Processes",
        }
        inst = wb["A_InstData"]
        assert inst["I20"].value == DEMO["installation_name_en"]
        assert inst["I21"].value == DEMO["street_number"]
        assert inst["I25"].value == "Ningbo"
        assert inst["I26"].value == "China"
        assert inst["E62"].value == "Iron or steel products"
        assert inst["E83"].value == "Iron or steel products"
        assert inst["L83"].value == "Bolt & fastener finishing line"

        products = wb["Summary_Products"]
        assert products["D10"].value == "Bolt & fastener finishing line"
        assert products["F10"].value == "73181588"
        assert products["H10"].value == "Hex bolt M12"
        assert float(products["I10"].value) == 1.87
        assert float(products["J10"].value) == 0.12
    finally:
        wb.close()
        out.unlink(missing_ok=True)


def test_download_endpoint():
    client = TestClient(app)
    res = client.post(
        "/api/routes/cbam-communication-xlsx/download",
        json={"workbook_values": DEMO},
    )
    assert res.status_code == 200
    assert (
        res.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "CBAM_Communication" in res.headers.get("content-disposition", "")
    assert len(res.content) > 100_000

    tmp = Path("storage") / "_test_cbam_out.xlsx"
    tmp.parent.mkdir(exist_ok=True)
    tmp.write_bytes(res.content)
    wb = load_workbook(tmp, data_only=False)
    try:
        assert wb["A_InstData"]["I20"].value == DEMO["installation_name_en"]
    finally:
        wb.close()
        tmp.unlink(missing_ok=True)
