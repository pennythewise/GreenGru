"""Fill the official EU CBAM Communication template (.xlsx) with workbook values.

Preserves the Commission template structure (sheets, formulas, styles).
Writes only input cells on A_InstData / Summary_Products; Summary_Communication
and Summary_Processes mostly pull from those via formulas when opened in Excel.

SEE (direct/indirect) on Summary_Products are formula cells in the blank template;
for a filled communication export we write the already-computed SEE numbers into
those cells so the declarant-facing product row shows intensity without requiring
the full D_Processes → InputOutput calculation chain.
"""

from __future__ import annotations

import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TEMPLATE_FILENAME = "CBAM-communication-template.xlsx"
OUTPUT_FILENAME = "CBAM_Communication_template_filled.xlsx"

# Official English label from CONST_LIST_Goods / Translations
DEFAULT_AGGREGATED_GOOD = "Iron or steel products"
DEFAULT_COUNTRY_NAME = "China"

# Field key → (sheet, cell). Keys match frontend cbam-workbook / localStorage.
FIELD_CELLS: dict[str, tuple[str, str]] = {
    # A_InstData — reporting period
    "period_start": ("A_InstData", "I9"),
    "period_end": ("A_InstData", "L9"),
    "reporting_period_start": ("A_InstData", "I9"),
    "reporting_period_end": ("A_InstData", "L9"),
    # A_InstData — installation
    "installation_name_optional": ("A_InstData", "I19"),
    "installation_name_en": ("A_InstData", "I20"),
    "street_number": ("A_InstData", "I21"),
    "economic_activity": ("A_InstData", "I22"),
    "post_code": ("A_InstData", "I23"),
    "po_box": ("A_InstData", "I24"),
    "city": ("A_InstData", "I25"),
    "country": ("A_InstData", "I26"),
    "unlocode": ("A_InstData", "I27"),
    "latitude": ("A_InstData", "I28"),
    "longitude": ("A_InstData", "I29"),
    "authorized_rep_name": ("A_InstData", "I30"),
    "authorized_rep_email": ("A_InstData", "I31"),
    "authorized_rep_telephone": ("A_InstData", "I32"),
    # A_InstData — verifier (optional)
    "verifier_company": ("A_InstData", "I37"),
    "verifier_street": ("A_InstData", "I38"),
    "verifier_city": ("A_InstData", "I39"),
    "verifier_postcode": ("A_InstData", "I40"),
    "verifier_country": ("A_InstData", "I41"),
    "verifier_rep_name": ("A_InstData", "I45"),
    "verifier_rep_email": ("A_InstData", "I46"),
    "verifier_rep_telephone": ("A_InstData", "I47"),
    "verifier_accreditation_body": ("A_InstData", "I48"),
    "verifier_registration_number": ("A_InstData", "I49"),
    # Summary_Products — first data row (row 10); row 9 is the Commission example
    "cn_code": ("Summary_Products", "F10"),
    "product_name": ("Summary_Products", "H10"),
    "see_direct": ("Summary_Products", "I10"),
    "see_indirect": ("Summary_Products", "J10"),
    "reducing_agent": ("Summary_Products", "P10"),
    "steel_mill_id": ("Summary_Products", "Q10"),
    "pct_mn": ("Summary_Products", "R10"),
    "pct_cr": ("Summary_Products", "S10"),
    "pct_ni": ("Summary_Products", "T10"),
    "pct_other_alloys": ("Summary_Products", "U10"),
    "pct_carbon": ("Summary_Products", "V10"),
    "scrap_per_t_steel": ("Summary_Products", "AB10"),
    "pct_other_materials": ("Summary_Products", "AC10"),
    "pct_pre_consumer_scrap": ("Summary_Products", "AD10"),
    "currency": ("Summary_Products", "AP10"),
    "carbon_price_due": ("Summary_Products", "AR10"),
}

COUNTRY_CODE_TO_NAME: dict[str, str] = {
    "CN": "China",
    "CHN": "China",
    "HK": "China, Hong Kong Special Administrative Region",
    "MO": "China, Macao Special Administrative Region",
    "TW": "Taiwan",
    "DE": "Germany",
    "FR": "France",
    "IT": "Italy",
    "ES": "Spain",
    "NL": "Netherlands",
    "BE": "Belgium",
    "PL": "Poland",
    "AT": "Austria",
    "SE": "Sweden",
    "FI": "Finland",
    "DK": "Denmark",
    "IE": "Ireland",
    "PT": "Portugal",
    "GR": "Greece",
    "CZ": "Czechia",
    "RO": "Romania",
    "HU": "Hungary",
    "SK": "Slovakia",
    "BG": "Bulgaria",
    "HR": "Croatia",
    "SI": "Slovenia",
    "LT": "Lithuania",
    "LV": "Latvia",
    "EE": "Estonia",
    "LU": "Luxembourg",
    "MT": "Malta",
    "CY": "Cyprus",
    "US": "United States",
    "JP": "Japan",
    "KR": "Korea, Republic of",
    "IN": "India",
    "VN": "Viet Nam",
    "TH": "Thailand",
    "MY": "Malaysia",
    "ID": "Indonesia",
    "SG": "Singapore",
    "AU": "Australia",
    "GB": "United Kingdom",
    "UK": "United Kingdom",
}


def template_path() -> Path:
    here = Path(__file__).resolve()
    candidates = [
        here.parents[1] / "static" / "templates" / TEMPLATE_FILENAME,
        here.parents[3] / "frontend" / "public" / "templates" / TEMPLATE_FILENAME,
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise FileNotFoundError(
        f"CBAM communication template not found. Expected one of: "
        f"{', '.join(str(c) for c in candidates)}"
    )


def _parse_date(value: str) -> date | str:
    text = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return text


def _parse_number(value: str) -> float | str:
    text = value.strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return value


def _normalize_cn_code(value: str) -> str:
    """CNCodes_ListKey uses SUBSTITUTE(spaces) — write digits-only key."""
    return "".join(ch for ch in value if ch.isdigit()) or value.strip()


def _normalize_country(value: str) -> str:
    text = value.strip()
    if not text:
        return text
    upper = text.upper()
    if upper in COUNTRY_CODE_TO_NAME:
        return COUNTRY_CODE_TO_NAME[upper]
    return text


def _coerce_value(key: str, raw: str) -> Any:
    if key in {
        "period_start",
        "period_end",
        "reporting_period_start",
        "reporting_period_end",
    }:
        return _parse_date(raw)
    if key == "country" or key == "verifier_country":
        return _normalize_country(raw)
    if key == "cn_code":
        return _normalize_cn_code(raw)
    if key in {
        "latitude",
        "longitude",
        "see_direct",
        "see_indirect",
        "see_total",
        "pct_mn",
        "pct_cr",
        "pct_ni",
        "pct_other_alloys",
        "pct_carbon",
        "scrap_per_t_steel",
        "pct_other_materials",
        "pct_pre_consumer_scrap",
        "carbon_price_due",
        "share_default_values",
        "embedded_electricity",
        "electricity_ef",
        "total_direct",
        "total_indirect",
        "total_emissions",
    }:
        return _parse_number(raw)
    return raw


def _set_cell(ws, addr: str, value: Any) -> None:
    ws[addr] = value


def fill_cbam_communication_xlsx(workbook_values: dict[str, Any]) -> Path:
    """Copy the official template, fill mapped cells, return path to temp .xlsx."""
    values = {
        str(k): str(v).strip()
        for k, v in (workbook_values or {}).items()
        if v is not None and str(v).strip() != ""
    }

    src = template_path()
    tmp = tempfile.NamedTemporaryFile(
        prefix="cbam-comm-",
        suffix=".xlsx",
        delete=False,
    )
    tmp_path = Path(tmp.name)
    tmp.close()
    shutil.copy2(src, tmp_path)

    wb = load_workbook(tmp_path)
    written: set[tuple[str, str]] = set()

    for key, (sheet, cell) in FIELD_CELLS.items():
        raw = values.get(key)
        if raw is None:
            continue
        if (sheet, cell) in written:
            continue
        if sheet not in wb.sheetnames:
            continue
        _set_cell(wb[sheet], cell, _coerce_value(key, raw))
        written.add((sheet, cell))

    # Aggregated goods + production process (drives Summary_* formula sheets)
    inst = wb["A_InstData"]
    goods = (
        values.get("aggregated_goods_category")
        or values.get("aggregated_good_type")
        or DEFAULT_AGGREGATED_GOOD
    )
    process_name = values.get("process_name") or values.get("production_process") or "P1"
    # Strip "P1 · " style prefixes for cleaner Name column; keep full string if short
    if "·" in process_name:
        process_name = process_name.split("·", 1)[-1].strip() or process_name

    if ("A_InstData", "E62") not in written:
        _set_cell(inst, "E62", goods)
    if ("A_InstData", "E83") not in written:
        _set_cell(inst, "E83", goods)
    if ("A_InstData", "F83") not in written:
        _set_cell(inst, "F83", goods)
    if ("A_InstData", "L83") not in written:
        _set_cell(inst, "L83", process_name)

    # Summary_Products process selector must match A_InstData!T83 (name or "P1 - goods")
    products = wb["Summary_Products"]
    if ("Summary_Products", "D10") not in written:
        _set_cell(products, "D10", process_name)

    # If country still looks like a code after generic pass missed it
    if not values.get("country"):
        _set_cell(inst, "I26", DEFAULT_COUNTRY_NAME)

    # see_total: keep SUM formula unless only total provided
    if "see_total" in values and "see_direct" not in values and "see_indirect" not in values:
        _set_cell(products, "K10", _coerce_value("see_total", values["see_total"]))

    wb.save(tmp_path)
    wb.close()
    return tmp_path
